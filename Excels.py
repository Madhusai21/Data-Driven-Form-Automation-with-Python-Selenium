import openpyxl
from openpyxl.styles import PatternFill
# file='C:/Users/yanna/Documents/EXCEL-1.xlsx'
def get_row_num(file,sheetname):
    work_book=openpyxl.load_workbook(file)
    sheet=work_book[sheetname]
    return sheet.max_row
def get_col_num(file,sheetname):
    work_book=openpyxl.load_workbook(file)
    sheet=work_book[sheetname]
    return sheet.max_column
# reading data
def read_cell_data(file,sheetname,rows,cols):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book[sheetname]
    return sheet.cell(row=rows,column=cols).value
#-----writing different data  ------
def write_cell_data(file,sheetname,rows,cols,data):
    # value=[
    #     ["ID", "Name", "Age", "City"],
    #     [1, "Madhu", 25, "Hyderabad"],
    #     [2, "Sai", 28, "Bangalore"],
    #     [3, "Om", 22, "Chennai"],
    #     [4, "Nikhil", 30, "Delhi"],
    #     [5, "Anjali", 27, "Mumbai"]
    # ]
    work_book = openpyxl.load_workbook(file)
    sheet = work_book[sheetname]
    sheet.cell(rows, cols).value = data
    # for rx in range(1,rows+1):
    #     for cx in range(1,cols+1):
    #          sheet.cell(rx,cx).value=data #-- by using list ....
            # sheet.cell(rx,cx).value=data[(rx,cx)] ---> by using dictionary ....
    work_book.save(file)

 # Explanation ++++
# sheet.cell(row=r, column=c).value = values[r - 1][c - 1]
# sheet.cell(row=r, column=c) → refers to the cell at row r and column c in the Excel sheet.
#.value = values[r - 1][c - 1] → assigns a value from the 2D list values to that cell.
# r - 1 and c - 1 are used because:
# Excel rows/columns start at 1.
# Python lists (like values) use zero-based indexing (start at 0).

# we can pass data by using dictionary also ...
# data = {
# (1,1): "ID", (1,2): "Name", (1,3): "Age", (1,4): "City",
#     (2,1): 1, (2,2): "Madhu", (2,3): 25, (2,4): "Hyderabad",
#     (3,1): 2, (3,2): "Sai", (3,3): 28, (3,4): "Bangalore",
#     (4,1): 3, (4,2): "Om", (4,3): 22, (4,4): "Chennai",
#     (5,1): 4, (5,2): "Nikhil", (5,3): 30, (5,4): "Delhi",
#     (6,1): 5, (6,2): "Anjali", (6,3): 27, (6,4): "Mumbai",
# }
def color_cells(file, sheetname , rows ,cols):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book['Sheet1']
    green_color = PatternFill(start_color='00ff00', fill_type='solid')
    sheet.cell(rows,cols).fill=green_color
    work_book.save(file)